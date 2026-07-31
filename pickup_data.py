"""
Datový model a export pro Pickup Report (Parkhotel TYCHO).

3 Month Pickup Update Report — sleduje výhled rezervací (Room Nights,
Revenue) na 3 měsíce dopředu oproti budgetu a loňskému roku.

Budget RN a Budget Rev. pro rok 2027 jsou převzaty z přílohy
"February 2027 PHT Daily.xls" (list "Start data", řádky
"Rooms Occupied" a "Rooms without breakfast)"). Pro ostatní roky se
budget musí doplnit ručně (nebo později napojit přes API hotelového PMS).
"""
import calendar
import json
from pathlib import Path

DATA_DIR = Path("data")
HOTEL_NAME = "Parkhotel TYCHO"
SELLABLE_ROOMS_DEFAULT = 22

MESICE = [
    "Leden", "Únor", "Březen", "Duben", "Květen", "Červen",
    "Červenec", "Srpen", "Září", "Říjen", "Listopad", "Prosinec",
]

# Budget RN (room nights) a Budget Rev. (tržby z ubytování) na rok 2027,
# zdroj: February 2027 PHT Daily.xls -> list "Start data".
SEED_BUDGET_2027 = {
    1:  {"budget_rn": 140.492,            "budget_rev": 166241.37376},
    2:  {"budget_rn": 154.73919999999998, "budget_rev": 190725.34835199997},
    3:  {"budget_rn": 147.51659999999998, "budget_rev": 210901.53268799998},
    4:  {"budget_rn": 176.41799999999998, "budget_rev": 278302.92336},
    5:  {"budget_rn": 205.0774,           "budget_rev": 353832.342864},
    6:  {"budget_rn": 217.932,            "budget_rev": 386750.84447999997},
    7:  {"budget_rn": 277.7104,           "budget_rev": 383351.43616},
    8:  {"budget_rn": 297.9658,           "budget_rev": 411311.99032},
    9:  {"budget_rn": 229.086,            "budget_rev": 395255.82096000004},
    10: {"budget_rn": 215.78480000000002,"budget_rev": 351038.71264000004},
    11: {"budget_rn": 164.142,            "budget_rev": 234670.53456},
    12: {"budget_rn": 124.3968,           "budget_rev": 159456.794112},
}


def empty_actual():
    return {"definite": {"yesterday": 0, "today": 0},
            "tentative": {"yesterday": 0, "today": 0}}


def build_month(year: int, month: int) -> dict:
    days_in_month = calendar.monthrange(year, month)[1]
    seed = SEED_BUDGET_2027.get(month, {}) if year == 2027 else {}
    return {
        "days_in_month": days_in_month,
        "sellable_rooms": SELLABLE_ROOMS_DEFAULT,
        "budget_rn": round(seed.get("budget_rn", 0), 2),
        "budget_rev": round(seed.get("budget_rev", 0), 2),
        "rn": empty_actual(),
        "revenue": empty_actual(),
        "last_year": {"rn": 0, "revenue": 0},
    }


def build_year(year: int) -> dict:
    return {
        "year": year,
        "hotel": HOTEL_NAME,
        "report_month": 2 if year == 2027 else 1,
        "months": {str(m): build_month(year, m) for m in range(1, 13)},
    }


def data_path(year: int) -> Path:
    return DATA_DIR / f"pickup_{year}.json"


def load_year(year: int) -> dict:
    p = data_path(year)
    if not p.exists():
        data = build_year(year)
        save_year(year, data)
        return data
    data = json.loads(p.read_text())
    # doplnit chybějící měsíce/klíče při rozšíření datového modelu
    changed = False
    for m in range(1, 13):
        if str(m) not in data.get("months", {}):
            data.setdefault("months", {})[str(m)] = build_month(year, m)
            changed = True
    if changed:
        save_year(year, data)
    return data


def save_year(year: int, data: dict):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    data_path(year).write_text(json.dumps(data, ensure_ascii=False, indent=2))


# ── Odvozené hodnoty (stejná logika jako na frontendu) ───────────────────────

def compute_month(m: dict) -> dict:
    inv = m["days_in_month"] * m["sellable_rooms"]

    def totals(block):
        d, t = block["definite"], block["tentative"]
        return {
            "definite":  {**d, "pu": d["today"] - d["yesterday"]},
            "tentative": {**t, "pu": t["today"] - t["yesterday"]},
            "total": {
                "yesterday": d["yesterday"] + t["yesterday"],
                "today":     d["today"] + t["today"],
                "pu":        (d["today"] + t["today"]) - (d["yesterday"] + t["yesterday"]),
            },
        }

    rn = totals(m["rn"])
    rev = totals(m["revenue"])

    ly_rn, ly_rev = m["last_year"]["rn"], m["last_year"]["revenue"]

    rn["budget"] = m["budget_rn"]
    rn["act_vs_budget"] = rn["total"]["today"] - m["budget_rn"]
    rn["last_year"] = ly_rn
    rn["act_vs_last_year"] = rn["total"]["today"] - ly_rn

    rev["budget"] = m["budget_rev"]
    rev["act_vs_budget"] = rev["total"]["today"] - m["budget_rev"]
    rev["last_year"] = ly_rev
    rev["act_vs_last_year"] = rev["total"]["today"] - ly_rev

    def safe_div(a, b):
        return a / b if b else 0

    def ratio_section(num, den, budget_num, budget_den, ly_num, ly_den):
        out = {}
        for row_key in ("definite", "tentative", "total"):
            y = safe_div(num[row_key]["yesterday"], den[row_key]["yesterday"])
            t = safe_div(num[row_key]["today"], den[row_key]["today"])
            out[row_key] = {"yesterday": y, "today": t, "pu": t - y}
        out["budget"] = safe_div(budget_num, budget_den)
        out["last_year"] = safe_div(ly_num, ly_den)
        out["act_vs_budget"] = out["total"]["today"] - out["budget"]
        out["act_vs_last_year"] = out["total"]["today"] - out["last_year"]
        return out

    occ = ratio_section(rn, {k: {"yesterday": inv, "today": inv} for k in ("definite", "tentative", "total")},
                         m["budget_rn"], inv, ly_rn, inv)
    adr = ratio_section(rev, rn, m["budget_rev"], m["budget_rn"], ly_rev, ly_rn)
    revpar = ratio_section(rev, {k: {"yesterday": inv, "today": inv} for k in ("definite", "tentative", "total")},
                            m["budget_rev"], inv, ly_rev, inv)

    return {
        "potential_inventory": inv,
        "rn": rn, "revenue": rev,
        "occupancy": occ, "adr": adr, "revpar": revpar,
    }


# ── Export do XLSX ───────────────────────────────────────────────────────────

def export_xlsx(year: int, data: dict, out_path: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    BLUE = Font(name="Arial", size=10, color="0000FF")
    BLUE_B = Font(name="Arial", size=10, color="0000FF", bold=True)
    BLACK = Font(name="Arial", size=10, color="000000")
    BLACK_B = Font(name="Arial", size=10, color="000000", bold=True)
    RED = Font(name="Arial", size=10, color="FF0000")
    RED_B = Font(name="Arial", size=10, color="FF0000", bold=True)
    HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
    HEAD_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    SECT_FILL = PatternFill("solid", fgColor="BDD7EE")
    THIN = Side(style="thin", color="B7B7B7")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def val_font(value, bold=False, editable=False):
        if value is not None and isinstance(value, (int, float)) and value < 0:
            return RED_B if bold else RED
        if editable:
            return BLUE_B if bold else BLUE
        return BLACK_B if bold else BLACK

    wb = Workbook()
    wb.remove(wb.active)

    # ── List: Přehled roku ──
    ws = wb.create_sheet("Přehled roku")
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"PICKUP REPORT — {data.get('hotel', HOTEL_NAME)} — ROK {year}"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1F4E79")
    headers = ["Měsíc", "RN Today", "Budget RN", "RN vs Budget", "LY RN", "RN vs LY",
               "Rev. Today", "Budget Rev.", "Rev. vs Budget", "LY Rev.", "Rev. vs LY",
               "Occ. %", "ADR", "RevPAR"]
    r0 = 3
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r0, column=c, value=h)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    row = r0 + 1
    sums = dict.fromkeys(["rn_today", "budget_rn", "ly_rn", "rev_today", "budget_rev", "ly_rev"], 0)
    for m in range(1, 13):
        md = data["months"][str(m)]
        c = compute_month(md)
        vals = [
            MESICE[m - 1],
            c["rn"]["total"]["today"], c["rn"]["budget"], c["rn"]["act_vs_budget"],
            c["rn"]["last_year"], c["rn"]["act_vs_last_year"],
            c["revenue"]["total"]["today"], c["revenue"]["budget"], c["revenue"]["act_vs_budget"],
            c["revenue"]["last_year"], c["revenue"]["act_vs_last_year"],
            c["occupancy"]["total"]["today"], c["adr"]["total"]["today"], c["revpar"]["total"]["today"],
        ]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.border = BORDER
            if ci == 1:
                cell.font = BLACK_B
                continue
            cell.font = val_font(v)
            if ci == 12:
                cell.number_format = "0.0%"
            elif ci in (7, 8, 9, 10, 11, 13, 14):
                cell.number_format = "#,##0"
            else:
                cell.number_format = "#,##0.0"
        sums["rn_today"] += c["rn"]["total"]["today"]
        sums["budget_rn"] += c["rn"]["budget"]
        sums["ly_rn"] += c["rn"]["last_year"]
        sums["rev_today"] += c["revenue"]["total"]["today"]
        sums["budget_rev"] += c["revenue"]["budget"]
        sums["ly_rev"] += c["revenue"]["last_year"]
        row += 1

    total_row = row
    ws.cell(row=total_row, column=1, value="CELKEM ROK").font = BLACK_B
    tot_vals = [
        sums["rn_today"], sums["budget_rn"], sums["rn_today"] - sums["budget_rn"],
        sums["ly_rn"], sums["rn_today"] - sums["ly_rn"],
        sums["rev_today"], sums["budget_rev"], sums["rev_today"] - sums["budget_rev"],
        sums["ly_rev"], sums["rev_today"] - sums["ly_rev"],
    ]
    for ci, v in enumerate(tot_vals, start=2):
        cell = ws.cell(row=total_row, column=ci, value=v)
        cell.font = val_font(v, bold=True)
        cell.border = BORDER
        cell.number_format = "#,##0" if ci >= 7 else "#,##0.0"

    widths = [12, 10, 10, 12, 8, 10, 12, 12, 13, 11, 11, 9, 9, 9]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── List na měsíc ──
    SECTIONS = [
        ("rn", "ROOM NIGHTS", "RN", "#,##0.0", False),
        ("occupancy", "OCCUPANCY", "%", "0.0%", True),
        ("adr", "ADR", "ADR", "#,##0", False),
        ("revenue", "REVENUE", "Rev.", "#,##0", False),
        ("revpar", "RevPAR", "RevPar", "#,##0", False),
    ]
    EDITABLE = {"rn", "revenue"}

    for m in range(1, 13):
        md = data["months"][str(m)]
        c = compute_month(md)
        ws = wb.create_sheet(f"{m:02d} {MESICE[m-1]}")
        ws.sheet_view.showGridLines = False
        ws["A1"] = f"{MESICE[m-1].upper()} {year} — 3 MONTH PICKUP UPDATE REPORT"
        ws["A1"].font = Font(name="Arial", size=13, bold=True, color="1F4E79")
        ws["A3"] = "Days in Month:"; ws["B3"] = md["days_in_month"]; ws["B3"].font = BLACK_B
        ws["C3"] = "Sellable Rooms:"; ws["D3"] = md["sellable_rooms"]; ws["D3"].font = BLUE_B
        ws["E3"] = "Potential Inventory:"; ws["F3"] = c["potential_inventory"]; ws["F3"].font = BLACK_B

        r = 5
        cols = ["", "Yesterday", "Today", "P/U", "Budget", "ACT vs Budget", "Last Year", "ACT vs Last Year"]
        for key, title, unit, numfmt, is_pct in SECTIONS:
            ws.cell(row=r, column=1, value=title).font = Font(name="Arial", bold=True, color="1F4E79", size=11)
            r += 1
            for ci, h in enumerate(cols, start=1):
                cell = ws.cell(row=r, column=ci, value=h)
                cell.font = HEAD_FONT
                cell.fill = SECT_FILL if ci > 0 else HEAD_FILL
                cell.font = Font(name="Arial", size=9, bold=True, color="1F4E79")
                cell.border = BORDER
            r += 1
            sec = c[key]
            editable = key in EDITABLE
            for row_label, row_key in [("Definite", "definite"), ("Tentative", "tentative"), ("TOTAL", "total")]:
                ws.cell(row=r, column=1, value=row_label).font = BLACK_B if row_key == "total" else BLACK
                rd = sec[row_key]
                y, t, pu = rd["yesterday"], rd["today"], rd["pu"]
                is_total = row_key == "total"
                y_f = val_font(y, editable=editable and not is_total)
                t_f = val_font(t, editable=editable and not is_total)
                for ci, v in [(2, y), (3, t), (4, pu)]:
                    cell = ws.cell(row=r, column=ci, value=v)
                    cell.font = val_font(v, editable=(editable and not is_total and ci in (2, 3)))
                    cell.number_format = numfmt
                    cell.border = BORDER
                if is_total:
                    for ci, v in [(5, sec["budget"]), (6, sec["act_vs_budget"]),
                                  (7, sec["last_year"]), (8, sec["act_vs_last_year"])]:
                        cell = ws.cell(row=r, column=ci, value=v)
                        cell.font = val_font(v)
                        cell.number_format = numfmt
                        cell.border = BORDER
                else:
                    for ci in (5, 6, 7, 8):
                        ws.cell(row=r, column=ci, value=None).border = BORDER
                r += 1
            r += 1

        for i, w in enumerate([14, 12, 12, 12, 12, 14, 12, 15], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
