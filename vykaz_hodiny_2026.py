"""
Generátor tabulky odpracovaných hodin pro rok 2026.
Výstup: vykaz_hodiny_2026.xlsx
"""
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── české státní svátky 2026 ──────────────────────────────────────────────────
HOLIDAYS = {
    date(2026, 1, 1):  "Nový rok",
    date(2026, 4, 3):  "Velký pátek",
    date(2026, 4, 6):  "Velikonoční pondělí",
    date(2026, 5, 1):  "Svátek práce",
    date(2026, 5, 8):  "Den vítězství",
    date(2026, 7, 5):  "Cyril a Metoděj",
    date(2026, 7, 6):  "Jan Hus",
    date(2026, 9, 28): "Den české státnosti",
    date(2026, 10, 28):"Den vzniku ČSR",
    date(2026, 11, 17):"Den boje za svobodu",
    date(2026, 12, 24):"Štědrý den",
    date(2026, 12, 25):"1. svátek vánoční",
    date(2026, 12, 26):"2. svátek vánoční",
}

MESICE = ["Leden","Únor","Březen","Duben","Květen","Červen",
          "Červenec","Srpen","Září","Říjen","Listopad","Prosinec"]

SAZBA = 400  # Kč/hod
HODINY_DEN = 8

def pracovni_dny(mesic: int) -> int:
    """Počet pracovních dní v daném měsíci (bez státních svátků)."""
    pocet = 0
    d = date(2026, mesic, 1)
    while d.month == mesic:
        if d.weekday() < 5 and d not in HOLIDAYS:
            pocet += 1
        d += timedelta(days=1)
    return pocet


def build_workbook() -> Workbook:
    wb = Workbook()

    # ── barvy ─────────────────────────────────────────────────────────────────
    COL_HEADER  = "1F4E79"   # tmavě modrá – záhlaví
    COL_SUBHEAD = "2E75B6"   # středně modrá – popis sloupce
    COL_INPUT   = "FFF2CC"   # žlutá – buňky pro vstup uživatele
    COL_CALC    = "DEEAF1"   # světle modrá – výpočtové buňky
    COL_TOTAL   = "BDD7EE"   # modrá – součtový řádek
    COL_WHITE   = "FFFFFF"
    COL_HOLIDAY = "FCE4D6"   # oranžová – svátky

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border_thin():
        s = Side(style="thin", color="AAAAAA")
        return Border(left=s, right=s, top=s, bottom=s)

    def border_medium():
        s = Side(style="medium", color="1F4E79")
        return Border(left=s, right=s, top=s, bottom=s)

    # ════════════════════════════════════════════════════════════════════════════
    # LIST 1 – Přehled roku
    # ════════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Přehled 2026"

    # Záhlaví
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "VÝKAZ PRÁCE – PŘEHLED 2026   |   Šimon Jeleňák"
    c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill = fill(COL_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Podtitul
    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = "Sazba: 400 Kč/hod  |  IČ: 43758207  |  Zdětín 151, 294 71 Benátky nad Jizerou"
    c.font = Font(name="Calibri", italic=True, size=10, color="FFFFFF")
    c.fill = fill(COL_SUBHEAD)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Záhlaví sloupců
    headers = [
        ("A3", "MĚSÍC"),
        ("B3", "Prac. dny"),
        ("C3", "Fond hod.\n(bez svátků)"),
        ("D3", "Dovolená\n(dny)"),
        ("E3", "Nemoc\n(dny)"),
        ("F3", "Odprac.\nhodiny"),
        ("G3", "Celkem\nKč"),
        ("H3", "Svátky v měsíci"),
    ]
    for cell_ref, label in headers:
        c = ws[cell_ref]
        c.value = label
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = fill(COL_SUBHEAD)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border_thin()
    ws.row_dimensions[3].height = 32

    # Datové řádky
    for m in range(1, 13):
        row = m + 3  # řádky 4–15
        pd = pracovni_dny(m)

        # svátky v tomto měsíci
        svat = [f"{d.strftime('%d.%m')} {name}"
                for d, name in HOLIDAYS.items() if d.month == m]
        svat_str = ", ".join(svat) if svat else "–"

        # A – Měsíc
        c = ws.cell(row, 1, MESICE[m - 1])
        c.font = Font(name="Calibri", bold=True, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = border_thin()

        # B – Pracovní dny (pevná hodnota)
        c = ws.cell(row, 2, pd)
        c.font = Font(name="Calibri", size=10)
        c.fill = fill(COL_WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_thin()

        # C – Fond hodin (pevný výpočet)
        c = ws.cell(row, 3, pd * HODINY_DEN)
        c.font = Font(name="Calibri", size=10)
        c.fill = fill(COL_CALC)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_thin()

        # D – Dovolená (vstup uživatele, počet DNÍ)
        c = ws.cell(row, 4, 0)
        c.font = Font(name="Calibri", size=10)
        c.fill = fill(COL_INPUT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_thin()

        # E – Nemoc (vstup uživatele, počet DNÍ)
        c = ws.cell(row, 5, 0)
        c.font = Font(name="Calibri", size=10)
        c.fill = fill(COL_INPUT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_thin()

        # F – Odpracované hodiny = (prac_dny - dovolená - nemoc) * 8
        col_c = f"C{row}"
        col_d = f"D{row}"
        col_e = f"E{row}"
        c = ws.cell(row, 6)
        c.value = f"=({col_c}-({col_d}+{col_e})*{HODINY_DEN})"
        c.font = Font(name="Calibri", bold=True, size=10)
        c.fill = fill(COL_CALC)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_thin()

        # G – Celkem Kč
        c = ws.cell(row, 7)
        c.value = f"=F{row}*{SAZBA}"
        c.font = Font(name="Calibri", size=10)
        c.fill = fill(COL_CALC)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = '#,##0 "Kč"'
        c.border = border_thin()

        # H – Svátky
        c = ws.cell(row, 8, svat_str)
        c.font = Font(name="Calibri", size=9, color="595959")
        c.fill = fill(COL_HOLIDAY) if svat else fill(COL_WHITE)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True, indent=1)
        c.border = border_thin()

        ws.row_dimensions[row].height = 18

    # Součtový řádek
    row_sum = 16
    ws.merge_cells(f"A{row_sum}:C{row_sum}")
    c = ws[f"A{row_sum}"]
    c.value = "CELKEM 2026"
    c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    c.fill = fill(COL_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_medium()
    ws.row_dimensions[row_sum].height = 22

    for col, formula in [(4, "=SUM(D4:D15)"), (5, "=SUM(E4:E15)"),
                         (6, "=SUM(F4:F15)"), (7, "=SUM(G4:G15)")]:
        c = ws.cell(row_sum, col)
        c.value = formula
        c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        c.fill = fill(COL_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_medium()
        if col == 7:
            c.number_format = '#,##0 "Kč"'

    # Prázdná buňka H součtu
    c = ws.cell(row_sum, 8)
    c.fill = fill(COL_HEADER)
    c.border = border_medium()

    # Legenda pod tabulkou
    ws.merge_cells("A18:H18")
    c = ws["A18"]
    c.value = ("💡  Žluté buňky (D, E) vyplňte ručně počtem dní dovolené / nemoci. "
               "Ostatní hodnoty se přepočítají automaticky.")
    c.font = Font(name="Calibri", italic=True, size=9, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A19:H19")
    c = ws["A19"]
    c.value = "Fond hodin = pracovní dny × 8 hod.  |  Odprac. hodiny = (prac. dny − dovolená − nemoc) × 8"
    c.font = Font(name="Calibri", italic=True, size=9, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center")

    # Šířky sloupců
    col_widths = [13, 11, 13, 12, 12, 14, 14, 45]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Zmrazit záhlaví
    ws.freeze_panes = "A4"

    # ════════════════════════════════════════════════════════════════════════════
    # LIST 2 – Státní svátky (přehled)
    # ════════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Státní svátky 2026")

    ws2.merge_cells("A1:C1")
    c = ws2["A1"]
    c.value = "STÁTNÍ SVÁTKY ČR 2026"
    c.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    c.fill = fill(COL_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 24

    for hdr, col in [("Datum", "A2"), ("Den", "B2"), ("Název", "C2")]:
        c = ws2[col]
        c.value = hdr
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = fill(COL_SUBHEAD)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_thin()

    DAYS_CZ = ["Pondělí","Úterý","Středa","Čtvrtek","Pátek","Sobota","Neděle"]
    for i, (d, name) in enumerate(sorted(HOLIDAYS.items()), start=3):
        ws2.cell(i, 1, d.strftime("%d.%m.%Y")).border = border_thin()
        ws2.cell(i, 1).alignment = Alignment(horizontal="center")
        ws2.cell(i, 2, DAYS_CZ[d.weekday()]).border = border_thin()
        ws2.cell(i, 2).alignment = Alignment(horizontal="center")
        ws2.cell(i, 3, name).border = border_thin()
        row_fill = fill(COL_HOLIDAY) if d.weekday() < 5 else fill(COL_WHITE)
        for col in range(1, 4):
            ws2.cell(i, col).fill = row_fill

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 28

    return wb


if __name__ == "__main__":
    wb = build_workbook()
    path = "vykaz_hodiny_2026.xlsx"
    wb.save(path)
    print(f"Soubor uložen: {path}")
